from copy import copy
import math
import json
import os
import re
import time
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter, range_boundaries

import pandas as pd
from rapidfuzz import fuzz
from ultralytics import YOLO

from express import OCR_PATTERN as OCR_PATTERNS

# ———RUTAS y UMBRALES ———
MODEL_PATH      = r'C:\Users\juans\Documents\proarchitecg\version_2_docker\model_clasification_image_v2\runs\classify\person_cls\weights\best.pt'
IMAGES_ROOT     = r'D:\historias\dev\imagenes_por_doc\LETRA A\ACTA N° 70\10'
OCR_ROOT        = r'D:\historias\dev\ocr_por_doc\LETRA A\ACTA N° 70\10'
OUTPUT_FILE     = r'C:\Users\juans\Documents\proarchitecg\version_2_docker\resultados_completos_v_final.xlsx'
TEMPLATE_PATH   = r'C:\Users\juans\Downloads\dev_prev\FORMATO HOJA DE CONTROL DOCUMENTAL.xlsx'
OUTPUT_DIR_CTRL = r'C:\Users\juans\Documents\proarchitecg\version_2_docker\model_clasification_image_v2\answer\hojas_control'

CONF_THRESH = 0.5
SCORING_MIN = 2
# ----------------- UTILIDADES -----------------
def extract_page_number(file_name: str) -> float:
    match = (re.search(r"pagina[_-]?(\d+)", file_name, re.IGNORECASE)
            or re.search(r"(\d+)", file_name))
    return int(match.group(1)) if match else math.nan

def compile_dict(raw_dict):
    compiled = {}
    for label, patterns in raw_dict.items():
        patterns = patterns if isinstance(patterns, list) else [patterns]
        compiled[label] = [
            pattern if isinstance(pattern, re.Pattern)
            else re.compile(rf"\b{pattern}\b", re.IGNORECASE | re.VERBOSE)
            for pattern in patterns
        ]
    return compiled


def fuzzy_ocr_label(text: str, patterns_by_label: dict, threshold: int = 75):
    best_label, best_score = "", 0.0
    for label, pattern_list in patterns_by_label.items():
        for pattern in pattern_list:
            score = fuzz.partial_ratio(text, pattern)
            if score > best_score:
                best_label, best_score = label, score
    if best_score >= threshold:
        return best_label, best_score / 100.0
    return "", 0.0


def visual_predict(model, image_path: str, strict: bool = True):
    result = model.predict(source=image_path, device="cpu", task="classify", verbose=False)[0]
    probabilities = getattr(result, "probs", None)
    arr = probabilities.data.tolist() if hasattr(probabilities, "data") else list(probabilities or [])
    if not arr or (strict and max(arr) < CONF_THRESH):
        return None, (max(arr) if arr else 0.0)
    index = arr.index(max(arr))
    return model.names[index], max(arr)

# ----------------- OCR / IMÁGENES -----------------
def normalize_filename(name: str) -> str:
    name = unicodedata.normalize("NFKD", name).encode("ASCII", "ignore").decode()
    return re.sub(r"\s+", " ", name).strip().lower()


def find_persona_images(root_path: str):
    persona_images = {}
    display_names = {}
    for dirpath, _, files in os.walk(root_path):
        images = [
            os.path.join(dirpath, file_name)
            for file_name in files
            if file_name.lower().endswith((".png", ".jpg", ".jpeg"))
        ]
        if images:
            raw_name = os.path.basename(dirpath)
            persona_name = normalize_text(raw_name)
            persona_images.setdefault(persona_name, []).extend(images)
            display_names[persona_name] = normalize_filename(raw_name)
    return persona_images, display_names


def build_ocr_map(root_path: str):
    json_map = {}
    if os.path.isfile(root_path):
        file_name = os.path.basename(root_path)
        if file_name.lower().endswith(".json"):
            normalized_name = normalize_text(os.path.splitext(file_name)[0])
            json_map[normalized_name] = root_path
        return json_map

    for dirpath, _, files in os.walk(root_path):
        for file_name in files:
            if file_name.lower().endswith(".json"):
                normalized_name = normalize_text(os.path.splitext(file_name)[0])
                json_map[normalized_name] = os.path.join(dirpath, file_name)
    return json_map


def match_json_for_persona(json_map, persona_name, threshold=70):
    if persona_name in json_map:
        return json_map[persona_name]

    best_score, best_key = 0, None
    for json_key in json_map:
        score = fuzz.partial_ratio(persona_name, json_key)
        if score > best_score:
            best_score, best_key = score, json_key
    if best_score >= threshold:
        return json_map[best_key]
    return None


OCR_REGEX = compile_dict(OCR_PATTERNS)


def classify(text, model, image_path):
    text = normalize_text(text or "")

    scores = {doc_type: 0 for doc_type in OCR_REGEX}
    for doc_type, pattern_list in OCR_REGEX.items():
        for pattern in pattern_list:
            if pattern.search(text):
                scores[doc_type] += 1

    thresholds = {doc_type: max(1, len(OCR_REGEX[doc_type]) // 2) for doc_type in OCR_REGEX}
    best_doc_type, count = max(scores.items(), key=lambda x: x[1])
    if count >= thresholds[best_doc_type]:
        return best_doc_type, count / len(OCR_REGEX[best_doc_type]), "ocr_scoring"

    for doc_type, pattern_list in OCR_REGEX.items():
        if any(pattern.search(text) for pattern in pattern_list):
            return doc_type, 1.0, "ocr_regex"

    simple = {doc_type: [pattern.pattern for pattern in OCR_REGEX[doc_type]] for doc_type in OCR_REGEX}
    label_fuzzy, confidence_fuzzy = fuzzy_ocr_label(text, simple)
    if label_fuzzy:
        return label_fuzzy, confidence_fuzzy, "ocr_fuzzy"

    label, confidence = visual_predict(model, image_path, strict=True)
    if label:
        return label, confidence, "visual"

    return "", 0.0, "none"
# ----------------- EXCEL -----------------
def copy_row_format(sheet, src_row, tgt_row, max_col=13, row_height=48):
    sheet.row_dimensions[tgt_row].height = row_height
    thin = Side(border_style="thin", color="000000")
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, max_col + 1):
        src = sheet.cell(row=src_row, column=col)
        tgt = sheet.cell(row=tgt_row, column=col)
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.fill = copy(src.fill)
            tgt.number_format = copy(src.number_format)
            tgt.protection = copy(src.protection)
            tgt.alignment = copy(src.alignment)
        tgt.border = full_border

    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row == src_row == merged.max_row:
            c1 = get_column_letter(merged.min_col)
            c2 = get_column_letter(merged.max_col)
            sheet.merge_cells(f"{c1}{tgt_row}:{c2}{tgt_row}")


def remove_holes(sheet, hole_ranges):
    parsed = []
    for rng in hole_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        parsed.append((min_row, max_row, min_col, max_col))

    to_unmerge = []
    for merged in list(sheet.merged_cells.ranges):
        for min_row, max_row, min_col, max_col in parsed:
            if not (merged.max_row < min_row or merged.min_row > max_row
                    or merged.max_col < min_col or merged.min_col > max_col):
                to_unmerge.append(merged.coord)
                break
    for coord in to_unmerge:
        sheet.unmerge_cells(coord)

    for rng in hole_ranges:
        for row in sheet[rng]:
            for cell in row:
                cell.value = None


def generate_control_sheet(persona_df, persona_name):
    output_path = os.path.join(OUTPUT_DIR_CTRL, f"{persona_name}_hoja_de_control.xlsx")
    if os.path.exists(output_path):
        print(f"⚠️ Ya existe hoja de control para '{persona_name}', omitiendo.")
        return

    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook.active
    START_ROW = 18

    holes = ["B57:B59", "C57:F59", "D60:F60"]
    remove_holes(sheet, holes)

    data_start = START_ROW
    data_end = START_ROW + len(persona_df) - 1
    for merged in list(sheet.merged_cells.ranges):
        if (
            merged.max_row >= data_start and merged.min_row <= data_end
            and merged.max_col >= 1 and merged.min_col <= 7
        ):
            sheet.unmerge_cells(str(merged))

    for index, record in enumerate(persona_df.sort_values("posicion").itertuples(), start=1):
        row = START_ROW + index - 1
        sheet.cell(row=row, column=1, value=index)                   # A
        sheet.cell(row=row, column=5, value=record.predicted)        # E
        sheet.cell(row=row, column=6, value=int(record.posicion))    # F
        sheet.cell(row=row, column=7, value=int(record.posicion))    # G

    workbook.save(output_path)
    print("✅ Control inmediato:", output_path)

# ----------------- MAIN -----------------
def main():
    os.makedirs(OUTPUT_DIR_CTRL, exist_ok=True)
    model = YOLO(MODEL_PATH)

    json_map = build_ocr_map(OCR_ROOT)
    persona_images, display_names = find_persona_images(IMAGES_ROOT)

    all_rows, global_id = [], 1

    for persona_key, image_paths in persona_images.items():
        display_name = display_names.get(persona_key, persona_key)
        print(f"\nProcesando persona '{display_name}' con {len(image_paths)} imágenes…")
        text_lookup = {}
        persona_records = []

        json_path = match_json_for_persona(json_map, persona_key)
        if json_path:
            with open(json_path, encoding="utf-8") as file:
                data = json.load(file)
            records = data if isinstance(data, list) else [data]
            for record in records:
                page = record.get("pagina")
                image_from_record = record.get("imagen", "")
                text = record.get("texto", "")
                if page is not None:
                    try:
                        num = int(page)
                        text_lookup[num] = text
                        text_lookup[str(num)] = text
                    except (ValueError, TypeError):
                        text_lookup[str(page)] = text
                if image_from_record:
                    text_lookup[os.path.basename(image_from_record)] = text

        for image_path in sorted(image_paths, key=extract_page_number):
            page = extract_page_number(image_path)
            basename = os.path.basename(image_path)
            text = (text_lookup.get(basename)
                    or text_lookup.get(page)
                    or text_lookup.get(str(page), ""))
            label, score, layer = classify(text, model, image_path)

            record = {
                "id":        global_id,
                "persona":   persona_key,
                "imagen":    image_path,
                "posicion":  page,
                "predicted": label,
                "score":     score,
                "layer":     layer,
            }
            all_rows.append(record)
            persona_records.append(record)
            global_id += 1

        if persona_records:
            persona_df = pd.DataFrame(persona_records)
            persona_df["correct"] = persona_df["persona"] == persona_df["predicted"]
            generate_control_sheet(persona_df, display_name)

    print("\n⏳ Generando DataFrame y guardando", OUTPUT_FILE)
    df = pd.DataFrame(all_rows)
    df["correct"] = df["persona"] == df["predicted"]
    df.to_excel(OUTPUT_FILE, index=False)
    print("✅ Consolidado en", OUTPUT_FILE)

    START_ROW = 18
    for persona_name, group in df[df["predicted"] != ""].groupby("persona"):
        workbook = load_workbook(TEMPLATE_PATH)
        sheet = workbook.active

        footer = None
        for row in sheet.iter_rows(min_row=START_ROW, max_row=sheet.max_row):
            for cell in row:
                if isinstance(cell.value, str) and "NOMBRE Y APELLIDOS" in cell.value.upper():
                    footer = cell.row
                    break
            if footer:
                break
        if not footer:
            footer = START_ROW + 38

        template_rows = footer - START_ROW
        num_pages = len(group)
        if num_pages > template_rows:
            extra = num_pages - template_rows
            sheet.insert_rows(footer, amount=extra)
            src_row = footer - 1
            for i in range(extra):
                dst_row = footer + i
                sheet.row_dimensions[dst_row].height = sheet.row_dimensions[src_row].height
                for col in (5, 6, 7):
                    src = sheet.cell(row=src_row, column=col)
                    dst = sheet.cell(row=dst_row, column=col)
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format

        for index, record in enumerate(group.sort_values("posicion").itertuples()):
            row = START_ROW + index
            sheet.cell(row=row, column=5, value=record.predicted)
            sheet.cell(row=row, column=6, value=int(record.posicion))
            sheet.cell(row=row, column=7, value=int(record.posicion))

        display_name = display_names.get(persona_name, persona_name)
        out = os.path.join(OUTPUT_DIR_CTRL, f"{display_name}_hoja_de_control.xlsx")
        workbook.save(out)
        print("✅ Control:", out)


if __name__ == "__main__":
    main()
